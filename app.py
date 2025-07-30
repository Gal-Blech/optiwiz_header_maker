# Optiwiz Header Maker - Web Application
#
# This script creates a simple web UI to translate Excel files into YAML.
#
# Required Libraries:
# - streamlit: To create the web application UI.
# - openpyxl: To read .xlsx files and access cell formatting.
#
# How to Install:
# pip install streamlit openpyxl
#
# How to Run:
# 1. Save this code as a Python file (e.g., app.py).
# 2. Open your terminal or command prompt.
# 3. Navigate to the directory where you saved the file.
# 4. Run the command: streamlit run app.py

import streamlit as st
from openpyxl import load_workbook
import io

# --- Helper Functions ---

def get_merged_range_obj(sheet, cell):
    """
    Checks if a cell is part of a merged range.
    If so, returns the MergeCellRange object. Otherwise, returns None.
    """
    for merged_cell_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cell_range:
            return merged_cell_range
    return None

def format_color_hex(argb_hex):
    """Converts openpyxl's ARGB hex to a standard #RRGGBB hex."""
    if isinstance(argb_hex, str) and len(argb_hex) == 8:
        return f"#{argb_hex[2:]}"
    return None

def has_border(cell):
    """Checks if a cell has any border style applied."""
    return (cell.border.left.style or cell.border.right.style or 
            cell.border.top.style or cell.border.bottom.style)

def get_border_color(cell):
    """Checks all four sides of a cell for a color and returns the first one found."""
    for side in ('left', 'right', 'top', 'bottom'):
        border_side = getattr(cell.border, side)
        if border_side and border_side.color and border_side.color.type == 'rgb':
            color = format_color_hex(border_side.color.rgb)
            if color and color.upper() != '#000000':
                return color
    return None

# --- Manual YAML Builder ---

def build_yaml_string(all_rows_data):
    """
    Manually builds the YAML string from the processed data to ensure
    the exact required output format, including special quoting rules.
    """
    lines = ["template:", "    format:", "        page_header:"]
    for row in all_rows_data:
        if not row:
            lines.append("            - []")
            continue
        
        lines.append("            -")
        for cell in row:
            if cell is None:
                lines.append("                - null")
                continue
            
            lines.append("                -")
            for key, value in cell.items():
                if key == 'merge':
                    lines.append("                    merge:")
                    lines.append(f"                        from_to: '{value['from_to']}'")
                elif key == 'type':
                    lines.append(f"                    {key}: {value}")
                elif key == 'value' and value == 'return "<placeholder>"':
                    lines.append(f"                    {key}: '{value}'")
                elif isinstance(value, bool):
                     lines.append(f"                    {key}: {str(value).lower()}")
                elif isinstance(value, (int, float)):
                     lines.append(f"                    {key}: {value}")
                else:
                     lines.append(f"                    {key}: '{value}'")
    
    lines.append("            - []")
    return "\n".join(lines)


# --- Core Translation Logic ---

def generate_yaml_from_file(file_object):
    """
    Reads an in-memory Excel file object, translates it, and returns the YAML as a string.
    Also returns any warnings generated during the process.
    """
    warnings = []
    all_rows_data = []
    
    workbook = load_workbook(file_object)
    sheet = workbook.active
    
    # Find the maximum column index that actually has content or styling
    max_col = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None or cell.has_style:
                max_col = max(max_col, cell.column)

    for row in sheet.iter_rows(max_col=max_col):
        row_data = []
        is_empty_row = all(cell.value is None and not cell.has_style for cell in row)

        if is_empty_row:
            all_rows_data.append([])
            continue

        for cell in row:
            merged_range_obj = get_merged_range_obj(sheet, cell)
            
            if merged_range_obj and cell.coordinate != merged_range_obj.coord.split(':')[0]:
                leader_cell = sheet[merged_range_obj.coord.split(':')[0]]
                if has_border(leader_cell):
                    row_data.append({'border': 1})
                else:
                    row_data.append(None)
                continue

            cell_obj = {}

            if merged_range_obj:
                cell_obj['merge'] = {'from_to': merged_range_obj.coord}

            value = cell.value
            if value is not None:
                if str(value).strip() == '<Logo>':
                    cell_obj['type'] = 'logo'
                    cell_obj['value'] = True
                    cell_below = sheet.cell(row=cell.row + 1, column=cell.column)
                    if cell_below.value:
                        warnings.append(f"**Logo Warning:** Data '{cell_below.value}' in cell {cell_below.coordinate} "
                                        f"may be obscured by the Logo in {cell.coordinate}.")
                elif str(value).strip() == '<placeholder>':
                    cell_obj['type'] = 'expert'
                    cell_obj['value'] = 'return "<placeholder>"'
                else:
                    cell_obj['value'] = value
            
            if cell.has_style:
                if cell.font.bold: cell_obj['bold'] = True
                if cell.font.name and cell.font.name.lower() != 'calibri': cell_obj['font_name'] = cell.font.name.lower()
                if cell.font.size and cell.font.size != 11: cell_obj['font_size'] = int(cell.font.size)
                
                if cell.font.color and cell.font.color.type == 'rgb':
                     font_color = format_color_hex(cell.font.color.rgb)
                     if font_color and font_color.upper() != '#000000':
                        cell_obj['font_color'] = font_color

                if cell.fill.fill_type == 'solid' and cell.fill.start_color.type == 'rgb':
                    bg_color = format_color_hex(cell.fill.start_color.rgb)
                    if bg_color and bg_color.upper() != '#FFFFFF':
                        cell_obj['bg_color'] = bg_color

                if cell.alignment.horizontal and cell.alignment.horizontal != 'left':
                    cell_obj['align'] = cell.alignment.horizontal
                if cell.alignment.vertical and cell.alignment.vertical != 'bottom':
                    cell_obj['valign'] = 'vcenter' if cell.alignment.vertical == 'center' else cell.alignment.vertical
                
                if has_border(cell):
                    cell_obj['border'] = 1
                    # **FIXED** Check all sides for border color
                    border_color = get_border_color(cell)
                    if border_color:
                        cell_obj['border_color'] = border_color

            if not cell_obj:
                row_data.append(None)
            else:
                row_data.append(cell_obj)
        
        all_rows_data.append(row_data)

    # Use the manual builder to generate the final string
    yaml_string = build_yaml_string(all_rows_data)
    
    return yaml_string, warnings

# --- Streamlit User Interface ---

st.set_page_config(page_title="Optiwiz Header Maker", layout="wide")

st.title("📄 Optiwiz Header Maker")
st.write("Upload your Excel design file to instantly translate it into Optiwiz YAML code.")

if 'yaml_output' not in st.session_state:
    st.session_state['yaml_output'] = ""
if 'file_name' not in st.session_state:
    st.session_state['file_name'] = ""

uploaded_file = st.file_uploader(
    "Choose an Excel file (.xlsx)",
    type="xlsx",
    accept_multiple_files=False
)

if uploaded_file is not None:
    st.success(f"File '{uploaded_file.name}' uploaded successfully!")

    if st.button("Translate to YAML", type="primary"):
        with st.spinner("Translating..."):
            yaml_output, warnings = generate_yaml_from_file(uploaded_file)
            
            for warning in warnings:
                st.warning(warning)
            
            st.session_state['yaml_output'] = yaml_output
            st.session_state['file_name'] = uploaded_file.name.rsplit('.', 1)[0] + ".yaml"

if st.session_state['yaml_output']:
    st.subheader("Generated YAML Code")
    st.code(st.session_state['yaml_output'], language='yaml')
    
    st.download_button(
        label="⬇️ Download YAML File",
        data=st.session_state['yaml_output'],
        file_name=st.session_state['file_name'],
        mime='text/yaml'
    )
